from pyexpat import model
import random
import json
from re import template
from sqlite3 import Cursor
from tempfile import tempdir
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
import pyodbc
from pydantic import BaseModel

class cell(BaseModel):
    module_color: str
    discipline: str
    term: str
    zet: float

color_pallete_num = 0;
color_pallete = [{
    1:"#61F4DE",
    2:"#5DE8D9",
    3:"#5ADBD5",
    4:"#56CFD0",
    5:"#52C3CC",
    6:"#4FB6C7",
    7:"#4BAAC2",
    8:"#479DBE",
    9:"#4491B9",
    10:"#4085B4",
    20:"#3C78B0",
    21:"#386CAB",
    22:"#3560A7",
    23:"#3153A2",
    24:"#2D479D",
    25:"#262E94",
    26:"#262E94"
},{
    1:"#001219",
    2:"#005f73",
    3:"#0a9396",
    4:"#94d2bd",
    5:"#e9d8a6",
    6:"#005f73",
    7:"#ee9b00",
    8:"#ca6702",
    9:"#bb3e03",
    10:"#ae2012",
    20:"#0a9396",
    21:"#ca6702",
    22:"#bb3e03",
    23:"#ae2012",
    24:"#ee9b00",
    25:"#9b2226",
    26:"#001219"
},
{
    1:"#D32F2F",
    2:"#FF4081",
    3:"#7B1FA2",
    4:"#7C4DFF",
    5:"#448AFF",
    6:"#303F9F",
    7:"#00BCD4",
    8:"#0288D1",
    9:"#8BC34A",
    10:"#388E3C",
    20:"#FFC107",
    21:"#FFEB3B",
    22:"#FF5722",
    23:"##9E9E9E",
    24:"#5D4037",
    25:"#9b2226",
    26:"#001219"
},]

# функция подключения к базе данных, на вход требует путь к базе данных возвращает курсор, который указывает на БД

def connect_to_DateBase(fullname_db):
    try:
        conn_string = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + fullname_db
        conn = pyodbc.connect(conn_string)
        cursor = conn.cursor()
        print("Connected To Database")
        return cursor
    except pyodbc.Error as e:
        print("Error in Connection", e)


def sort_modul(date):
    buf = "Первый семестр"
    full_data = []
    for i in range(len(date)):
        date_dist = date[i]
        if date_dist[2] != buf:
            full_data += sorted(date[len(full_data):i])
            buf = date_dist[2]
    full_data += sorted(date[len(full_data):i + 1])
    return full_data


# функция делает запрос в базу данных и выводит нужные значения для дальнейшего вывода в карту
# (мудуль, дисциплина, семестр, зеты(складывая все за одну дисц)), на выходу лист из листов в каждом из которых находятся данные
def select_to_DataBase(cur):
    data_set = []
    sem = ["Первый", "Второй", "Третий", "Четвертый", "Пятый", "Шестой", "Седьмой", "Восьмой", ]
    data = []
    buf = ""
    zet = 0.0
    j = -1
    for i in range(len(sem)):
        cur.execute('SELECT ID_of_module, Discipline, Control_period, ZET  FROM Disciplines_and_practices WHERE Control_period LIKE ? AND ID_of_the_educational_program = 2', (sem[i] + " семестр"))
        for row in cur.fetchall():
            if buf != row[1]:
                buf = row[1]
                data.append(row[0])
                data.append(row[1])
                data.append(row[2])
                data_set.append(data.copy())
                data_rev = data_set[j]
                if len(data_rev) == 3:
                    data_rev.append(int(zet))
                    data_set[j] = data_rev.copy()
                else:
                    data_rev[3] = int(zet)
                    data_set[j] = data_rev.copy()
                zet = 0.0
                j += 1
            if row[3] != None:
                zet += round(float(row[3]), 1)

            data.clear()
    data_rev = data_set[-1]
    data_rev.append(int(zet))
    data_set[-1] = data_rev.copy()

    data_set = sort_modul(data_set)
    
    return data_set


def select_color(cur, modul):
    cur.execute(
        'SELECT Color  FROM Module_reference WHERE ID_of_module LIKE ?', (modul))
    for row in cur.fetchall():
        return (row[0])


def getTable():
    fullname_db = 'main\db.accdb'
    sem = ["Первый", "Второй", "Третий", "Четвертый", "Пятый", "Шестой", "Седьмой", "Восьмой", ]
    cur = connect_to_DateBase(fullname_db=fullname_db)
    data = select_to_DataBase(cur)
    cell_list = []

    for el in data:
        с = cell(module_color=color_pallete[color_pallete_num][el[0]],discipline=el[1],term=el[2],zet=el[3])
        cell_list.append(с)

    table = []
    for i in range(0,8):
        temp_list = []

        for el in cell_list:
            if el.term == sem[i] + ' семестр':
                temp_list.append(el.dict())

        table.append(temp_list)

    # print(table)
    for el in table[0]:
        print(el["discipline"])

    return table


# функция создает карту и задаем все данные кроме предметов в семестрах, на вход требует имя карты
def CreateMap(filename_map):
    wk = xlsxwriter.Workbook(filename_map)
    ws = wk.add_worksheet()
    ws.set_column(1, 29, 29)
    wk.close()
    workbook = openpyxl.load_workbook(filename_map)
    worksheet = workbook.active
    ns = NamedStyle(name='standart')
    ns.font = Font(bold=False, size=12)
    border = Side(style='thick', color='000000')
    ns.border = Border(left=border, top=border, right=border, bottom=border)
    ns.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    workbook.add_named_style(ns)
    worksheet.column_dimensions['A'].height = 50
    worksheet.row_dimensions[1].height = 50
    worksheet.merge_cells('A1:I1')
    worksheet['A1'] = 'КАРТА ДИСЦИПЛИН'
    worksheet['A1'].style = 'standart'
    worksheet['A1'].font = Font(bold=True, size=12)
    worksheet["A2"] = "З.Е."
    worksheet['A2'].style = 'standart'
    for col in range(3, 33):
        worksheet["A" + str(col)] = col - 2
        worksheet["A" + str(col)].style = 'standart'
    for col in range(ord('B'), ord('J')):
        worksheet[chr(col) + str(2)] = str(col - 65) + " семестр"
        worksheet[chr(col) + str(2)].style = 'standart'
    return worksheet, workbook


# заполняем данные, размер и цвет  в ячейках карты,
# Так же мы красим предметы в соответствии с модулем
def filling_map(fullname_db, filename_map):
    cur = connect_to_DateBase(fullname_db)
    date = select_to_DataBase(cur)
    
    # with open("json.json", "w", encoding='utf-8') as file:
    #     json.dump(date, file, ensure_ascii=False)

    # print(json.dumps(date).encode("utf-8"))

    ws, wk = CreateMap(filename_map)
    adr_cell = "B"
    buf = "Первый семестр"
    row = 3
    i = -1
    while adr_cell != "J" and i < len(date) - 1:
        i += 1
        date_dist = date[i]
        if date_dist[2] == buf and date_dist[3] != 0:
            dip = adr_cell + str(row) + ':' + adr_cell + str(row + date_dist[3] - 1)
            ws[adr_cell + str(row)].style = 'standart'
            ws[adr_cell + str(row)] = date_dist[1]
            cell = ws[adr_cell + str(row)]
            color = select_color(cur, date_dist[0])
            cell.fill = openpyxl.styles.PatternFill(start_color=str(color), end_color=str(color), fill_type='solid')
            ws.merge_cells(dip)
            row += date_dist[3] - 1
            buf = date_dist[2]
            row += 1
        elif date_dist[3] != 0:
            adr_cell = chr(ord(adr_cell) + 1)
            buf = date_dist[2]
            row = 3
            i -= 1
    wk.save(filename=filename_map)


# основная функция-связующая все части и вводит основные параметры всего
def main():
    filename_map = 'map.xlsx'
    fullname_db = 'main\db.accdb'
    # filling_map(fullname_db, filename_map)
    getTable();
    


if __name__ == "__main__":
    main()
